"""Экспорт отчётов в PDF, Excel (.xlsx) и CSV."""

from __future__ import annotations

import csv
import io
import os
from typing import Optional

from database import Database

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False


_THIN = Side(style='thin')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
_HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
_CELL_FONT = Font(size=10)
_CENTER = Alignment(horizontal='center', vertical='center')
_LEFT = Alignment(horizontal='left', vertical='center')


def _ensure_db(db: Optional[Database]) -> Database:
    return db if db is not None else Database()


# ──────────────────────────────── PDF ────────────────────────────────


def _build_pdf(title: str, headers: list[str], rows: list[list[str]]) -> bytes:
    if not HAS_REPORTLAB:
        raise RuntimeError("reportlab not installed")
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles['Title']), Spacer(1, 6*mm)]
    data = [headers] + rows
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#D9E2F3')]),
    ]))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()


def export_grades_pdf(subject_id: int, db: Optional[Database] = None) -> bytes:
    """Gradebook as PDF, returns bytes."""
    db = _ensure_db(db)
    summary = db.subject_summary(subject_id)
    subj_name = dict(summary)['name'] if summary else f'Предмет #{subject_id}'
    students, lessons, grades = db.subject_gradebook(subject_id)
    headers = ['Студент'] + [str(i + 1) for i in range(len(lessons))]
    rows = []
    for s in students:
        row = [f"{s['last_name']} {s['first_name']}"]
        for l in lessons:
            row.append(grades.get(str(s['id']), {}).get(str(l['id']), ''))
        rows.append(row)
    return _build_pdf(f'Ведомость: {subj_name}', headers, rows)


def export_report_pdf(date: str, db: Optional[Database] = None) -> bytes:
    """Daily report as PDF, returns bytes."""
    db = _ensure_db(db)
    report = db.daily_report(date)
    headers = ['Предмет', 'Группа', 'Статус', 'Оценок', 'Пропусков']
    rows = []
    for r in report:
        rd = dict(r)
        rows.append([
            rd.get('subject_name', ''),
            rd.get('group_name', ''),
            rd.get('status', ''),
            str(rd.get('grades_count', 0)),
            str(rd.get('absent_count', 0)),
        ])
    return _build_pdf(f'Отчёт за {date}', headers, rows)


# ──────────────────────────────── Excel ────────────────────────────────


def _build_xlsx(title: str, headers: list[str], rows: list[list[str]]) -> bytes:
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl not installed")
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    header_row = 3
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=ci, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
        c.border = _BORDER
    for ri, row in enumerate(rows, header_row + 1):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = _CELL_FONT
            c.alignment = _CENTER if ci > 1 else _LEFT
            c.border = _BORDER
    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 30)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_grades_xlsx(subject_id: int, db: Optional[Database] = None) -> bytes:
    """Gradebook as Excel, returns bytes."""
    db = _ensure_db(db)
    summary = db.subject_summary(subject_id)
    subj_name = dict(summary)['name'] if summary else f'Предмет #{subject_id}'
    students, lessons, grades = db.subject_gradebook(subject_id)
    headers = ['Студент'] + [str(i + 1) for i in range(len(lessons))]
    rows = []
    for s in students:
        row = [f"{s['last_name']} {s['first_name']}"]
        for l in lessons:
            row.append(grades.get(str(s['id']), {}).get(str(l['id']), ''))
        rows.append(row)
    return _build_xlsx(f'Ведомость: {subj_name}', headers, rows)


def export_report_xlsx(date: str, db: Optional[Database] = None) -> bytes:
    """Daily report as Excel, returns bytes."""
    db = _ensure_db(db)
    report = db.daily_report(date)
    headers = ['Предмет', 'Группа', 'Статус', 'Оценок', 'Пропусков']
    rows = []
    for r in report:
        rd = dict(r)
        rows.append([
            rd.get('subject_name', ''),
            rd.get('group_name', ''),
            rd.get('status', ''),
            str(rd.get('grades_count', 0)),
            str(rd.get('absent_count', 0)),
        ])
    return _build_xlsx(f'Отчёт за {date}', headers, rows)


# ──────────────────────────────── CSV (legacy) ────────────────────────────────


def export_grades_csv(subject_id: int, filepath: Optional[str] = None, db: Optional[Database] = None) -> str:
    """Экспортирует ведомость по предмету в CSV (таблица студенты × занятия)."""
    if db is None:
        db = Database()
    if filepath is None:
        filepath = os.path.join(os.path.dirname(__file__), f'grades_{subject_id}.csv')
    students, lessons, grades = db.subject_gradebook(subject_id)
    summary = db.subject_summary(subject_id)
    subj_name = dict(summary)['name'] if summary else f'Subject #{subject_id}'

    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow([subj_name])
        header = ['Студент'] + [str(i + 1) for i in range(len(lessons))]
        writer.writerow(header)
        for s in students:
            row = [f"{s['last_name']} {s['first_name']}"]
            for l in lessons:
                row.append(grades.get(str(s['id']), {}).get(str(l['id']), ''))
            writer.writerow(row)
    return filepath


def export_report_csv(date: str, filepath: Optional[str] = None, db: Optional[Database] = None) -> str:
    """Экспортирует дневной отчёт в CSV."""
    if db is None:
        db = Database()
    if filepath is None:
        filepath = os.path.join(os.path.dirname(__file__), f'report_{date}.csv')
    rows = db.daily_report(date)

    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow([f'Отчёт за {date}'])
        writer.writerow(['Предмет', 'Группа', 'Статус', 'Оценок', 'Пропусков'])
        for r in rows:
            rd = dict(r)
            writer.writerow([
                rd.get('subject_name', ''),
                rd.get('group_name', ''),
                rd.get('status', ''),
                rd.get('grades_count', 0),
                rd.get('absent_count', 0),
            ])
    return filepath
